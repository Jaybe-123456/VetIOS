import argparse
import json
import logging
import re
import zipfile
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from xml.etree import ElementTree

try:
    from import_clinical_cases import read_rows, transform_row
    from prepare_dataset import DATE_PATTERN, DIRECT_IDENTIFIER_PATTERNS
except ImportError:  # pragma: no cover - supports package-style execution
    from vetios_pipeline.import_clinical_cases import read_rows, transform_row
    from vetios_pipeline.prepare_dataset import DATE_PATTERN, DIRECT_IDENTIFIER_PATTERNS

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

TABULAR_EXTENSIONS = {".csv", ".jsonl", ".xlsx"}
TEXT_EXTENSIONS = {".txt", ".md", ".rst"}
DOCUMENT_EXTENSIONS = {".pdf", ".docx", ".pptx"}
SUPPORTED_EXTENSIONS = TABULAR_EXTENSIONS | TEXT_EXTENSIONS | DOCUMENT_EXTENSIONS | {".json"}


def find_identifier_hits(text: str, strict_dates: bool) -> List[str]:
    hits = [name for name, pattern in DIRECT_IDENTIFIER_PATTERNS.items() if pattern.search(text)]
    if strict_dates and DATE_PATTERN.search(text):
        hits.append("date")
    return hits


def iter_files(roots: List[Path]) -> Iterable[Path]:
    for root in roots:
        if not root.exists():
            raise FileNotFoundError(root)
        if root.is_file():
            yield root
            continue
        for path in root.rglob("*"):
            if path.is_file() and not path.name.startswith("."):
                yield path


def read_text(path: Path, max_chars: int) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")[:max_chars]


def read_pdf_text(path: Path, max_chars: int) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError("PDF extraction requires `pip install pypdf`") from exc

    reader = PdfReader(str(path))
    parts: List[str] = []
    for page in reader.pages:
        parts.append(page.extract_text() or "")
        if sum(len(part) for part in parts) >= max_chars:
            break
    return "\n".join(parts)[:max_chars]


def read_docx_text(path: Path, max_chars: int) -> str:
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    texts = [node.text or "" for node in root.findall(".//w:t", namespace)]
    return "\n".join(texts)[:max_chars]


def read_pptx_text(path: Path, max_chars: int) -> str:
    texts: List[str] = []
    namespace = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    with zipfile.ZipFile(path) as archive:
        slide_names = sorted(name for name in archive.namelist() if name.startswith("ppt/slides/slide"))
        for slide_name in slide_names:
            root = ElementTree.fromstring(archive.read(slide_name))
            texts.extend(node.text or "" for node in root.findall(".//a:t", namespace))
            if sum(len(text) for text in texts) >= max_chars:
                break
    return "\n".join(texts)[:max_chars]


def extract_document_text(path: Path, max_chars: int) -> str:
    suffix = path.suffix.lower()
    if suffix in TEXT_EXTENSIONS:
        return read_text(path, max_chars)
    if suffix == ".pdf":
        return read_pdf_text(path, max_chars)
    if suffix == ".docx":
        return read_docx_text(path, max_chars)
    if suffix == ".pptx":
        return read_pptx_text(path, max_chars)
    if suffix == ".json":
        return read_text(path, max_chars)
    raise ValueError(f"unsupported document extension: {suffix}")


def iter_xlsx_rows(path: Path) -> Iterable[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX import requires `pip install openpyxl`") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return
    keys = [str(header).strip() if header is not None else "" for header in headers]
    for values in rows:
        row = {
            key: value
            for key, value in zip(keys, values)
            if key and value is not None and str(value).strip()
        }
        if row:
            yield row


def iter_tabular_rows(path: Path) -> Iterable[Dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        yield from iter_xlsx_rows(path)
        return
    if suffix == ".csv":
        yield from read_rows(path, "csv")
        return
    if suffix == ".jsonl":
        yield from read_rows(path, "jsonl")
        return
    raise ValueError(f"unsupported tabular extension: {suffix}")


def importer_args(args: argparse.Namespace) -> argparse.Namespace:
    return argparse.Namespace(
        source=args.source,
        usage_class=args.usage_class,
        review_status=args.review_status,
        default_split=args.default_split,
        strict_dates=args.strict_dates,
        redact_direct_identifiers=args.redact_direct_identifiers,
        generate_case_id=args.generate_case_id,
        allow_unreviewed=args.allow_unreviewed,
        allow_unlabeled=args.allow_unlabeled,
        skip_unlabeled=args.skip_unlabeled,
        allow_duplicate_case_id=args.allow_duplicate_case_id,
    )


def import_tabular_file(path: Path, args: argparse.Namespace) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    imported: List[Dict[str, Any]] = []
    errors: List[str] = []
    warnings: List[str] = []
    base_args = importer_args(args)

    for row_index, row in enumerate(iter_tabular_rows(path), start=1):
        row_args = argparse.Namespace(**vars(base_args))
        if not row_args.source:
            row_args.source = f"{args.source_prefix}:{path.name}" if args.source_prefix else path.name
        try:
            transformed, row_warnings = transform_row(row, row_args)
        except Exception as exc:  # noqa: BLE001 - report all row-level import failures
            errors.append(f"row {row_index}: {exc}")
            continue
        warnings.extend(f"row {row_index}: {warning}" for warning in row_warnings)
        if transformed is not None:
            imported.append(transformed)

    status = "imported" if imported else "no_importable_rows"
    if errors and not imported:
        status = "failed"
    return imported, {
        "path": str(path),
        "kind": "tabular",
        "status": status,
        "imported_rows": len(imported),
        "error_count": len(errors),
        "warning_count": len(warnings),
        "errors": errors[:20],
        "warnings": warnings[:20],
    }


def chunk_text(text: str, chunk_chars: int) -> Iterable[str]:
    normalized = re.sub(r"\n{3,}", "\n\n", text).strip()
    for start in range(0, len(normalized), chunk_chars):
        chunk = normalized[start : start + chunk_chars].strip()
        if chunk:
            yield chunk


def process_document_file(path: Path, args: argparse.Namespace) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    try:
        text = extract_document_text(path, args.max_text_chars)
    except Exception as exc:  # noqa: BLE001 - keep inventory moving
        return [], {
            "path": str(path),
            "kind": "document",
            "status": "failed",
            "error": str(exc),
        }

    hits = find_identifier_hits(text, strict_dates=args.strict_dates)
    if hits and not args.include_flagged_documents:
        return [], {
            "path": str(path),
            "kind": "document",
            "status": "flagged_possible_phi",
            "identifier_hits": hits,
            "text_chars": len(text),
        }

    chunks = [
        {
            "source_path": str(path),
            "source_name": path.name,
            "source_type": path.suffix.lower().lstrip("."),
            "usage_class": args.usage_class,
            "text": chunk,
            "possible_identifier_hits": hits,
        }
        for chunk in chunk_text(text, args.chunk_chars)
    ]
    return chunks, {
        "path": str(path),
        "kind": "document",
        "status": "chunked" if chunks else "empty_text",
        "chunks": len(chunks),
        "text_chars": len(text),
        "identifier_hits": hits,
    }


def write_jsonl(path: Path, rows: List[Dict[str, Any]], append: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    mode = "a" if append else "w"
    with path.open(mode, encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def ingest_sources(args: argparse.Namespace) -> None:
    roots = [Path(root) for root in args.root]
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    inventory: List[Dict[str, Any]] = []
    clinical_rows: List[Dict[str, Any]] = []
    document_chunks: List[Dict[str, Any]] = []
    seen_case_ids = set()

    for path in iter_files(roots):
        suffix = path.suffix.lower()
        if suffix not in SUPPORTED_EXTENSIONS:
            inventory.append({"path": str(path), "kind": "unsupported", "status": "skipped"})
            continue

        if suffix in TABULAR_EXTENSIONS:
            rows, record = import_tabular_file(path, args)
            deduped_rows: List[Dict[str, Any]] = []
            for row in rows:
                case_id = row["case_id"]
                if case_id in seen_case_ids and not args.allow_duplicate_case_id:
                    record.setdefault("warnings", []).append(f"duplicate case_id skipped: {case_id}")
                    record["warning_count"] = record.get("warning_count", 0) + 1
                    continue
                seen_case_ids.add(case_id)
                deduped_rows.append(row)
            clinical_rows.extend(deduped_rows)
            record["imported_rows_after_dedupe"] = len(deduped_rows)
            inventory.append(record)
            continue

        chunks, record = process_document_file(path, args)
        document_chunks.extend(chunks)
        inventory.append(record)

    inventory_path = output_dir / "drive_inventory.json"
    documents_path = output_dir / "source_document_chunks.jsonl"
    clinical_output = Path(args.clinical_output)

    inventory_path.write_text(json.dumps(inventory, indent=2, ensure_ascii=False), encoding="utf-8")
    if document_chunks:
        write_jsonl(documents_path, document_chunks)
    if clinical_rows:
        write_jsonl(clinical_output, clinical_rows, append=args.append_clinical)

    logger.info("inventory written: %s", inventory_path)
    logger.info("clinical rows imported: %s", len(clinical_rows))
    if clinical_rows:
        logger.info("clinical output: %s", clinical_output)
    logger.info("document chunks written: %s", len(document_chunks))
    if document_chunks:
        logger.info("document chunk output: %s", documents_path)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Audit mounted/downloaded Google Drive sources and import only structured clinical case exports "
            "into VetIOS clinical_cases.jsonl."
        )
    )
    parser.add_argument("--root", action="append", required=True, help="Mounted/downloaded Drive folder path")
    parser.add_argument("--output-dir", default="vetios_pipeline/datasets/drive_audit")
    parser.add_argument("--clinical-output", default="vetios_pipeline/datasets/clinical_cases.jsonl")
    parser.add_argument("--source", default="", help="Default source for tabular rows")
    parser.add_argument("--source-prefix", default="drive", help="Source prefix when --source is absent")
    parser.add_argument(
        "--usage-class",
        default="internal_deidentified",
        choices=[
            "public_open",
            "public_restricted",
            "credentialed_deidentified",
            "internal_deidentified",
            "consented_research",
        ],
    )
    parser.add_argument("--review-status", default="clinician_reviewed")
    parser.add_argument("--default-split", default="train", choices=["train", "eval", "validation", "test"])
    parser.add_argument("--strict-dates", action="store_true")
    parser.add_argument("--redact-direct-identifiers", action="store_true")
    parser.add_argument("--generate-case-id", action="store_true")
    parser.add_argument("--allow-unreviewed", action="store_true")
    parser.add_argument("--allow-unlabeled", action="store_true")
    parser.add_argument("--skip-unlabeled", action="store_true")
    parser.add_argument("--allow-duplicate-case-id", action="store_true")
    parser.add_argument("--append-clinical", action="store_true")
    parser.add_argument("--include-flagged-documents", action="store_true")
    parser.add_argument("--max-text-chars", type=int, default=250000)
    parser.add_argument("--chunk-chars", type=int, default=2500)
    return parser


if __name__ == "__main__":
    ingest_sources(build_parser().parse_args())
